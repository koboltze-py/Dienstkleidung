"""
DRK Dienstkleidung - Bestandsverwaltung
Anzeige, Bearbeitung und Verwaltung des Kleidungsbestands.
"""

import os
from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QGridLayout, QLabel, QPushButton,
    QTableWidget, QTableWidgetItem, QComboBox, QLineEdit,
    QSpinBox, QDialog, QDialogButtonBox, QFormLayout, QFrame,
    QMessageBox, QHeaderView, QAbstractItemView, QSplitter,
    QScrollArea, QDateEdit, QCheckBox, QSizePolicy,
)
from PySide6.QtCore import Qt, QDate
from PySide6.QtGui import QColor, QFont

from utils import today_iso, format_datum


class BestandDialog(QDialog):
    """Dialog zum Hinzufügen/Bearbeiten eines Bestandseintrags."""

    def __init__(self, db, art_id: int = None, item: dict = None, parent=None):
        super().__init__(parent)
        self.db = db
        self.edit_item = item
        self.setWindowTitle("Bestandseintrag bearbeiten" if item else "Neuen Eintrag anlegen")
        self.setMinimumWidth(380)
        self.setModal(True)
        self._setup_ui(art_id)

    def _setup_ui(self, art_id):
        layout = QFormLayout(self)
        layout.setSpacing(12)
        layout.setContentsMargins(20, 20, 20, 20)

        # Kleidungsart
        self.cb_art = QComboBox()
        for art in self.db.get_kleidungsarten():
            self.cb_art.addItem(art["name"], art["id"])
        if art_id:
            idx = self.cb_art.findData(art_id)
            if idx >= 0:
                self.cb_art.setCurrentIndex(idx)
        layout.addRow("Kleidungsart:", self.cb_art)

        # Größe
        self.le_groesse = QLineEdit()
        self.le_groesse.setPlaceholderText("z.B. M, XL, 42, 38/40 ...")
        layout.addRow("Größe:", self.le_groesse)

        # Menge
        self.sb_menge = QSpinBox()
        self.sb_menge.setRange(0, 9999)
        layout.addRow("Menge auf Lager:", self.sb_menge)

        # Mindestbestand
        self.sb_min = QSpinBox()
        self.sb_min.setRange(0, 9999)
        self.sb_min.setToolTip("Bei Mindestbestand-Unterschreitung erscheint eine Warnung.")
        layout.addRow("Mindestbestand:", self.sb_min)

        # Bemerkung
        self.le_bem = QLineEdit()
        self.le_bem.setPlaceholderText("Optional")
        layout.addRow("Bemerkung:", self.le_bem)

        # Vordefinierte Werte übernehmen
        if self.edit_item:
            self.le_groesse.setText(str(self.edit_item.get("groesse", "")))
            self.sb_menge.setValue(int(self.edit_item.get("menge", 0)))
            self.sb_min.setValue(int(self.edit_item.get("min_menge", 0)))
            self.le_bem.setText(self.edit_item.get("bemerkung", ""))
            idx = self.cb_art.findData(self.edit_item.get("art_id"))
            if idx >= 0:
                self.cb_art.setCurrentIndex(idx)
            self.cb_art.setEnabled(False)
            self.le_groesse.setEnabled(False)

        buttons = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel
        )
        buttons.button(QDialogButtonBox.StandardButton.Ok).setText("Speichern")
        buttons.button(QDialogButtonBox.StandardButton.Cancel).setText("Abbrechen")
        buttons.accepted.connect(self._validate_and_accept)
        buttons.rejected.connect(self.reject)
        layout.addRow(buttons)

    def _validate_and_accept(self):
        if not self.le_groesse.text().strip():
            QMessageBox.warning(self, "Pflichtfeld", "Bitte eine Größe eingeben.")
            return
        self.accept()

    def get_data(self) -> dict:
        return {
            "art_id": self.cb_art.currentData(),
            "art_name": self.cb_art.currentText(),
            "groesse": self.le_groesse.text().strip(),
            "menge": self.sb_menge.value(),
            "min_menge": self.sb_min.value(),
            "bemerkung": self.le_bem.text().strip(),
        }


class EingangDialog(QDialog):
    """Dialog: Wareneingang – mehrere Positionen auf einmal buchen."""

    def __init__(self, db, art_id: int = None, groesse: str = "", parent=None):
        super().__init__(parent)
        self.db = db
        self.setWindowTitle("Wareneingang buchen")
        self.setMinimumWidth(680)
        self.setModal(True)
        self._rows: list[dict] = []
        self._setup_ui(art_id, groesse)

    def _setup_ui(self, art_id, groesse):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 16)
        layout.setSpacing(12)

        layout.addWidget(QLabel("Mehrere Positionen können gleichzeitig eingebucht werden."))

        # Datum + Bemerkung oben
        meta = QFormLayout()
        meta.setSpacing(8)
        self.de_datum = QDateEdit(QDate.currentDate())
        self.de_datum.setDisplayFormat("dd.MM.yyyy")
        self.de_datum.setCalendarPopup(True)
        self.de_datum.setMaximumWidth(150)
        meta.addRow("Datum:", self.de_datum)
        self.le_bem = QLineEdit()
        self.le_bem.setPlaceholderText("Lieferschein, Anmerkung ... (für alle Positionen)")
        meta.addRow("Bemerkung:", self.le_bem)
        layout.addLayout(meta)

        sep = QFrame(); sep.setFrameShape(QFrame.Shape.HLine); sep.setStyleSheet("color:#ddd;")
        layout.addWidget(sep)

        # Spaltenköpfe
        hdr_w = QWidget()
        hdr = QHBoxLayout(hdr_w)
        hdr.setContentsMargins(0, 0, 0, 0)
        hdr.setSpacing(8)
        for txt, w in [("Kleidungsart", 180), ("Größe", 120), ("Anzahl", 80)]:
            lb = QLabel(txt)
            lb.setFixedWidth(w)
            lb.setStyleSheet("font-weight:bold; color:#444;")
            hdr.addWidget(lb)
        hdr.addStretch()
        layout.addWidget(hdr_w)

        # Zeilen-Container
        self._rows_layout = QVBoxLayout()
        self._rows_layout.setSpacing(4)
        container_w = QWidget()
        container_w.setLayout(self._rows_layout)
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.Shape.NoFrame)
        scroll.setMaximumHeight(240)
        scroll.setWidget(container_w)
        layout.addWidget(scroll)

        btn_add = QPushButton("+ Zeile hinzufügen")
        btn_add.setObjectName("btn_secondary")
        btn_add.clicked.connect(lambda: self._add_zeile())
        layout.addWidget(btn_add)

        btns = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel
        )
        btns.button(QDialogButtonBox.StandardButton.Ok).setText("Einbuchen")
        btns.button(QDialogButtonBox.StandardButton.Cancel).setText("Abbrechen")
        btns.accepted.connect(self._validate_and_accept)
        btns.rejected.connect(self.reject)
        layout.addWidget(btns)

        # Erste Zeile mit Vorbelegung
        self._add_zeile(art_id, groesse)

    def _add_zeile(self, art_id=None, groesse=""):
        row_w = QWidget()
        hl = QHBoxLayout(row_w)
        hl.setContentsMargins(0, 0, 0, 0)
        hl.setSpacing(8)

        cb_art = QComboBox()
        cb_art.setFixedWidth(180)
        for art in self.db.get_kleidungsarten():
            cb_art.addItem(art["name"], art["id"])
        if art_id:
            idx = cb_art.findData(art_id)
            if idx >= 0:
                cb_art.setCurrentIndex(idx)

        le_gr = QLineEdit(groesse)
        le_gr.setPlaceholderText("z.B. M, 42 ...")
        le_gr.setFixedWidth(120)

        sb = QSpinBox()
        sb.setRange(1, 9999)
        sb.setValue(1)
        sb.setFixedWidth(80)

        btn_x = QPushButton("×")
        btn_x.setFixedSize(26, 26)
        btn_x.setStyleSheet("QPushButton{color:#B20000;font-weight:bold;font-size:15px;border:none;background:transparent;}"
                            "QPushButton:hover{background:#fce4e4;border-radius:4px;}")

        hl.addWidget(cb_art)
        hl.addWidget(le_gr)
        hl.addWidget(sb)
        hl.addWidget(btn_x)
        hl.addStretch()

        ri = {"widget": row_w, "cb_art": cb_art, "le_gr": le_gr, "sb": sb}
        btn_x.clicked.connect(lambda _, rw=row_w, r=ri: self._remove_zeile(rw, r))
        self._rows.append(ri)
        self._rows_layout.addWidget(row_w)

    def _remove_zeile(self, row_w, ri):
        if len(self._rows) <= 1:
            return
        self._rows.remove(ri)
        self._rows_layout.removeWidget(row_w)
        row_w.deleteLater()

    def _validate_and_accept(self):
        for ri in self._rows:
            if ri["le_gr"].text().strip():
                self.accept()
                return
        QMessageBox.warning(self, "Pflichtfeld", "Bitte mindestens eine Größe eingeben.")

    def get_data(self) -> list[dict]:
        datum = self.de_datum.date().toString("yyyy-MM-dd")
        bem = self.le_bem.text().strip()
        result = []
        for ri in self._rows:
            gr = ri["le_gr"].text().strip()
            if not gr:
                continue
            result.append({
                "art_id": ri["cb_art"].currentData(),
                "art_name": ri["cb_art"].currentText(),
                "groesse": gr,
                "menge": ri["sb"].value(),
                "datum": datum,
                "bemerkung": bem,
            })
        return result


class AusbuchenDialog(QDialog):
    """Dialog: Ware aus dem Bestand ausbuchen (Defekt, Verlust etc.)."""

    GRUENDE = ["Defekt", "Verlust", "Fehlbuchung", "Entsorgung", "Sonstiges"]

    def __init__(self, db, item: dict, parent=None):
        super().__init__(parent)
        self.db = db
        self._item = item
        self.setWindowTitle("Ware ausbuchen")
        self.setMinimumWidth(380)
        self.setModal(True)
        self._setup_ui()

    def _setup_ui(self):
        form = QFormLayout(self)
        form.setSpacing(12)
        form.setContentsMargins(20, 20, 20, 16)

        art = f"{self._item.get('art_name','')}  Gr. {self._item.get('groesse','')}"
        lbl = QLabel(art)
        lbl.setStyleSheet("font-weight:bold;")
        form.addRow("Artikel:", lbl)

        max_m = int(self._item.get("menge", 1))
        lbl_lager = QLabel(f"{max_m} Stk. auf Lager")
        form.addRow("Bestand:", lbl_lager)

        self.sb_menge = QSpinBox()
        self.sb_menge.setRange(1, max_m)
        self.sb_menge.setValue(1)
        form.addRow("Anzahl ausbuchen:", self.sb_menge)

        self.cb_grund = QComboBox()
        for g in self.GRUENDE:
            self.cb_grund.addItem(g)
        form.addRow("Grund:", self.cb_grund)

        self.le_bem = QLineEdit()
        self.le_bem.setPlaceholderText("Optional ...")
        form.addRow("Bemerkung:", self.le_bem)

        self.de_datum = QDateEdit(QDate.currentDate())
        self.de_datum.setDisplayFormat("dd.MM.yyyy")
        self.de_datum.setCalendarPopup(True)
        self.de_datum.setMaximumWidth(150)
        form.addRow("Datum:", self.de_datum)

        btns = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel
        )
        btns.button(QDialogButtonBox.StandardButton.Ok).setText("Ausbuchen")
        btns.button(QDialogButtonBox.StandardButton.Cancel).setText("Abbrechen")
        btns.accepted.connect(self.accept)
        btns.rejected.connect(self.reject)
        form.addRow(btns)

    def get_data(self) -> dict:
        return {
            "menge": self.sb_menge.value(),
            "grund": self.cb_grund.currentText(),
            "bemerkung": self.le_bem.text().strip(),
            "datum": self.de_datum.date().toString("yyyy-MM-dd"),
        }


class MultiAusbuchenDialog(QDialog):
    """Dialog: Mehrere Bestandspositionen auf einmal ausbuchen."""

    GRUENDE = ["Defekt", "Verlust", "Fehlbuchung", "Entsorgung", "Sonstiges"]

    def __init__(self, db, parent=None):
        super().__init__(parent)
        self.db = db
        self.setWindowTitle("Bestand ausbuchen")
        self.setMinimumWidth(780)
        self.setMinimumHeight(520)
        self.setModal(True)
        self._rows: list[dict] = []
        self._setup_ui()

    def _setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 16, 20, 16)
        layout.setSpacing(10)

        layout.addWidget(QLabel("Häkchen setzen und Anzahl wählen – alle markierten Positionen werden ausgebucht."))

        # Filter-Zeile
        filter_row = QHBoxLayout()
        filter_row.setSpacing(8)
        self._le_filter = QLineEdit()
        self._le_filter.setPlaceholderText("🔍  Filtern nach Kleidungsart oder Größe ...")
        self._le_filter.textChanged.connect(self._apply_filter)
        filter_row.addWidget(self._le_filter)
        layout.addLayout(filter_row)

        # Spaltenköpfe
        hdr_w = QWidget()
        hdr = QHBoxLayout(hdr_w)
        hdr.setContentsMargins(4, 0, 4, 0)
        hdr.setSpacing(8)
        for txt, w in [("✓", 28), ("Kleidungsart", 180), ("Größe", 90), ("Auf Lager", 80), ("Ausbuchen", 90)]:
            lb = QLabel(txt)
            lb.setFixedWidth(w)
            lb.setStyleSheet("font-weight:bold; color:#444;")
            hdr.addWidget(lb)
        hdr.addStretch()
        layout.addWidget(hdr_w)

        # Zeilen-Container in ScrollArea
        self._container_w = QWidget()
        self._rows_layout = QVBoxLayout(self._container_w)
        self._rows_layout.setContentsMargins(4, 4, 4, 4)
        self._rows_layout.setSpacing(3)
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.Shape.NoFrame)
        scroll.setWidget(self._container_w)
        layout.addWidget(scroll)

        # Globale Felder
        sep = QFrame(); sep.setFrameShape(QFrame.Shape.HLine); sep.setStyleSheet("color:#ddd;")
        layout.addWidget(sep)

        meta = QFormLayout()
        meta.setSpacing(8)
        self.cb_grund = QComboBox()
        for g in self.GRUENDE:
            self.cb_grund.addItem(g)
        meta.addRow("Grund (für alle):", self.cb_grund)
        self.le_bem = QLineEdit()
        self.le_bem.setPlaceholderText("Optional ...")
        meta.addRow("Bemerkung:", self.le_bem)
        self.de_datum = QDateEdit(QDate.currentDate())
        self.de_datum.setDisplayFormat("dd.MM.yyyy")
        self.de_datum.setCalendarPopup(True)
        self.de_datum.setMaximumWidth(140)
        meta.addRow("Datum:", self.de_datum)
        layout.addLayout(meta)

        btns = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel
        )
        btns.button(QDialogButtonBox.StandardButton.Ok).setText("Ausbuchen")
        btns.button(QDialogButtonBox.StandardButton.Cancel).setText("Abbrechen")
        btns.accepted.connect(self._validate_and_accept)
        btns.rejected.connect(self.reject)
        layout.addWidget(btns)

        self._load_items()

    def _load_items(self):
        self._all_items = self.db.get_bestand()
        self._apply_filter()

    def _apply_filter(self):
        text = self._le_filter.text().lower()
        filtered = [
            it for it in self._all_items
            if not text
            or text in it.get("art_name", "").lower()
            or text in str(it.get("groesse", "")).lower()
        ]
        # Bestehende Zeilen entfernen
        for ri in self._rows:
            ri["widget"].deleteLater()
        self._rows.clear()
        for item in filtered:
            self._add_row(item)

    def _add_row(self, item: dict):
        menge = int(item.get("menge", 0))
        row_w = QWidget()
        hl = QHBoxLayout(row_w)
        hl.setContentsMargins(0, 1, 0, 1)
        hl.setSpacing(8)

        chk = QCheckBox()
        chk.setFixedWidth(28)

        lbl_art = QLabel(item.get("art_name", ""))
        lbl_art.setFixedWidth(180)
        lbl_gr = QLabel(str(item.get("groesse", "")))
        lbl_gr.setFixedWidth(90)
        lbl_lager = QLabel(str(menge))
        lbl_lager.setFixedWidth(80)
        lbl_lager.setAlignment(Qt.AlignmentFlag.AlignCenter)
        if menge == 0:
            lbl_lager.setStyleSheet("color:#B20000; font-weight:bold;")

        sb = QSpinBox()
        sb.setRange(1, max(1, menge))
        sb.setValue(1)
        sb.setFixedWidth(90)
        sb.setEnabled(False)
        chk.toggled.connect(sb.setEnabled)

        hl.addWidget(chk)
        hl.addWidget(lbl_art)
        hl.addWidget(lbl_gr)
        hl.addWidget(lbl_lager)
        hl.addWidget(sb)
        hl.addStretch()

        self._rows.append({"widget": row_w, "chk": chk, "sb": sb, "item": item})
        self._rows_layout.addWidget(row_w)

    def _validate_and_accept(self):
        selected = [ri for ri in self._rows if ri["chk"].isChecked()]
        if not selected:
            QMessageBox.warning(self, "Keine Auswahl", "Bitte mindestens eine Position auswählen.")
            return
        self.accept()

    def get_selected(self) -> list[dict]:
        datum = self.de_datum.date().toString("yyyy-MM-dd")
        grund = self.cb_grund.currentText()
        bem = self.le_bem.text().strip()
        return [
            {
                "item": ri["item"],
                "menge": ri["sb"].value(),
                "datum": datum,
                "grund": grund,
                "bemerkung": bem,
            }
            for ri in self._rows
            if ri["chk"].isChecked()
        ]


class BestandView(QWidget):
    """Hauptansicht für die Bestandsverwaltung."""

    def __init__(self, db, parent=None):
        super().__init__(parent)
        self.db = db
        self._all_data: list[dict] = []
        self._block_tables: list = []
        self._setup_ui()
        self._load_data()

    def _setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(24, 20, 24, 20)
        layout.setSpacing(12)

        # Titel
        lbl_title = QLabel("Bestandsverwaltung")
        lbl_title.setObjectName("page_title")
        lbl_sub = QLabel("Kleidungsbestand anzeigen, bearbeiten und einbuchen")
        lbl_sub.setObjectName("page_subtitle")
        layout.addWidget(lbl_title)
        layout.addWidget(lbl_sub)

        # Toolbar
        toolbar = QHBoxLayout()
        toolbar.setSpacing(10)

        self.cb_filter = QComboBox()
        self.cb_filter.setMinimumWidth(180)
        self.cb_filter.addItem("Alle Kleidungsarten", None)
        self.cb_filter.currentIndexChanged.connect(self._apply_filter)
        toolbar.addWidget(QLabel("Filtern:"))
        toolbar.addWidget(self.cb_filter)

        self.le_search = QLineEdit()
        self.le_search.setPlaceholderText("Größe suchen ...")
        self.le_search.setMinimumWidth(160)
        self.le_search.textChanged.connect(self._apply_filter)
        toolbar.addWidget(self.le_search)

        toolbar.addStretch()

        btn_eingang = QPushButton("📦  Wareneingang")
        btn_eingang.setObjectName("btn_secondary")
        btn_eingang.clicked.connect(self._open_eingang)
        toolbar.addWidget(btn_eingang)

        btn_aus = QPushButton("➖  Ausbuchen")
        btn_aus.setObjectName("btn_secondary")
        btn_aus.setStyleSheet("color:#B20000;")
        btn_aus.clicked.connect(self._open_ausbuchen)
        toolbar.addWidget(btn_aus)

        btn_add = QPushButton("➕  Neues Kleidungsstück")
        btn_add.setObjectName("btn_primary")
        btn_add.clicked.connect(self._open_add)
        toolbar.addWidget(btn_add)

        btn_kat = QPushButton("🏷  Neue Kategorie")
        btn_kat.setObjectName("btn_secondary")
        btn_kat.clicked.connect(self._open_neue_kategorie)
        toolbar.addWidget(btn_kat)

        btn_exp = QPushButton("📊  Excel-Export")
        btn_exp.setObjectName("btn_secondary")
        btn_exp.clicked.connect(self._export_excel)
        toolbar.addWidget(btn_exp)

        layout.addLayout(toolbar)

        # Splitter: obere Kategorie-Blöcke + untere Detailansicht
        splitter = QSplitter(Qt.Orientation.Vertical)
        splitter.setChildrenCollapsible(True)

        # Kategorie-Blöcke in ScrollArea
        self._blocks_scroll = QScrollArea()
        self._blocks_scroll.setWidgetResizable(True)
        self._blocks_scroll.setFrameShape(QFrame.Shape.NoFrame)
        self._blocks_container = QWidget()
        self._blocks_layout = QGridLayout(self._blocks_container)
        self._blocks_layout.setSpacing(12)
        self._blocks_layout.setContentsMargins(0, 0, 8, 0)
        self._blocks_layout.setColumnStretch(0, 1)
        self._blocks_layout.setColumnStretch(1, 1)
        self._blocks_scroll.setWidget(self._blocks_container)
        splitter.addWidget(self._blocks_scroll)

        # Detailbereich: Buchungshistorie
        detail_frame = QFrame()
        detail_frame.setObjectName("stat_card")
        detail_layout = QVBoxLayout(detail_frame)
        detail_layout.setContentsMargins(12, 10, 12, 10)
        detail_layout.setSpacing(6)
        lbl_detail = QLabel("Buchungshistorie  — Zeile auswählen")
        bold_f = QFont(); bold_f.setBold(True)
        lbl_detail.setFont(bold_f)
        self._lbl_detail_title = lbl_detail
        detail_layout.addWidget(lbl_detail)

        self._tbl_buchungen = QTableWidget(0, 6)
        self._tbl_buchungen.setHorizontalHeaderLabels(
            ["Datum", "Typ", "Menge", "Mitarbeiter", "Ausgegeben von", "Bemerkung"]
        )
        self._tbl_buchungen.horizontalHeader().setStretchLastSection(True)
        self._tbl_buchungen.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        self._tbl_buchungen.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        self._tbl_buchungen.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        self._tbl_buchungen.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)
        self._tbl_buchungen.horizontalHeader().setSectionResizeMode(4, QHeaderView.ResizeMode.ResizeToContents)
        self._tbl_buchungen.verticalHeader().setVisible(False)
        self._tbl_buchungen.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self._tbl_buchungen.setAlternatingRowColors(True)
        detail_layout.addWidget(self._tbl_buchungen)
        splitter.addWidget(detail_frame)

        splitter.setSizes([500, 200])
        layout.addWidget(splitter)

        # Zusammenfassung
        self._lbl_summary = QLabel("Lade ...")
        self._lbl_summary.setObjectName("page_subtitle")
        layout.addWidget(self._lbl_summary)

    def _populate_filter(self):
        self.cb_filter.blockSignals(True)
        current = self.cb_filter.currentData()
        self.cb_filter.clear()
        self.cb_filter.addItem("Alle Kleidungsarten", None)
        for art in self.db.get_kleidungsarten():
            self.cb_filter.addItem(art["name"], art["id"])
        idx = self.cb_filter.findData(current)
        self.cb_filter.setCurrentIndex(max(0, idx))
        self.cb_filter.blockSignals(False)

    def _load_data(self):
        self._all_data = self.db.get_bestand()
        self._populate_filter()
        self._apply_filter()

    def _apply_filter(self):
        art_id = self.cb_filter.currentData()
        suche = self.le_search.text().strip().lower()

        filtered = [
            d for d in self._all_data
            if (art_id is None or d["art_id"] == art_id)
            and (not suche or suche in str(d["groesse"]).lower()
                 or suche in str(d["art_name"]).lower())
        ]
        self._fill_blocks(filtered)

    def _fill_blocks(self, data: list[dict]):
        # Vorhandene Blöcke leeren
        self._block_tables.clear()
        while self._blocks_layout.count():
            child = self._blocks_layout.takeAt(0)
            if child.widget():
                child.widget().deleteLater()

        # Nach Kleidungsart gruppieren (Reihenfolge erhalten)
        groups: dict[int, list[dict]] = {}
        for item in data:
            aid = item["art_id"]
            if aid not in groups:
                groups[aid] = []
            groups[aid].append(item)

        warn_color = QColor("#FFF3CD")
        grid_row = 0
        grid_col = 0

        for art_id, items in groups.items():
            art_name = items[0].get("art_name", "")
            total_art = sum(int(i.get("menge", 0)) for i in items)

            # --- Kategorie-Block ---
            block = QFrame()
            block.setObjectName("stat_card")
            block.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Preferred)
            bl = QVBoxLayout(block)
            bl.setContentsMargins(14, 10, 14, 10)
            bl.setSpacing(6)

            # Kopfzeile
            hdr = QHBoxLayout()
            lbl_name = QLabel(art_name)
            name_font = QFont()
            name_font.setBold(True)
            name_font.setPointSize(11)
            lbl_name.setFont(name_font)
            lbl_name.setStyleSheet("color: #B20000;")
            hdr.addWidget(lbl_name)
            hdr.addStretch()
            lbl_total = QLabel(f"Gesamt: {total_art} Stück")
            lbl_total.setStyleSheet("color: #666; font-size: 11px;")
            hdr.addWidget(lbl_total)
            btn_hinzu = QPushButton("➕ Größe hinzufügen")
            btn_hinzu.setObjectName("btn_secondary")
            btn_hinzu.clicked.connect(lambda chk, aid=art_id: self._open_add_for_art(aid))
            hdr.addWidget(btn_hinzu)
            bl.addLayout(hdr)

            # Mini-Tabelle für diese Kategorie
            tbl = QTableWidget(len(items), 4)
            tbl.setHorizontalHeaderLabels(["Größe", "Auf Lager", "Mindestbestand", "Bemerkung"])
            tbl.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
            tbl.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
            tbl.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
            tbl.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeMode.Stretch)
            tbl.verticalHeader().setVisible(False)
            tbl.verticalHeader().setDefaultSectionSize(34)
            tbl.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
            tbl.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
            tbl.setAlternatingRowColors(True)
            tbl.setShowGrid(True)
            tbl.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
            tbl.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)

            for r, item in enumerate(items):
                menge = int(item.get("menge", 0))
                min_m = int(item.get("min_menge", 0))
                is_low = min_m > 0 and menge <= min_m

                cells = [
                    str(item.get("groesse", "")),
                    str(menge),
                    str(min_m) if min_m > 0 else "–",
                    item.get("bemerkung", ""),
                ]
                for c, text in enumerate(cells):
                    it = QTableWidgetItem(text)
                    it.setData(Qt.ItemDataRole.UserRole, item)
                    if is_low:
                        it.setBackground(warn_color)
                    tbl.setItem(r, c, it)

            tbl.currentItemChanged.connect(self._on_row_selected)
            tbl.cellDoubleClicked.connect(self._on_cell_double_clicked)
            tbl.setToolTip("Doppelklick für Aktionen (Eingang, Bearbeiten, Löschen)")
            # Höhe exakt auf alle Zeilen setzen (kein Scrollbalken nötig)
            row_h = tbl.verticalHeader().defaultSectionSize()
            hdr_h = tbl.horizontalHeader().height()
            tbl.setFixedHeight(hdr_h + row_h * len(items) + 4)
            bl.addWidget(tbl)
            # Wrapper: füllt volle Spaltenbreite, Block bleibt oben
            wrapper = QWidget()
            wrapper.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Preferred)
            wl = QVBoxLayout(wrapper)
            wl.setContentsMargins(0, 0, 0, 0)
            wl.setSpacing(0)
            wl.addWidget(block)
            wl.addStretch(1)
            self._blocks_layout.addWidget(wrapper, grid_row, grid_col)
            grid_col += 1
            if grid_col > 1:
                grid_col = 0
                grid_row += 1

        # letzte Zeile: Stretch-Spacer damit Blöcke oben bleiben
        self._blocks_layout.setRowStretch(grid_row + 1, 1)

        total_all = sum(int(d.get("menge", 0)) for d in data)
        self._lbl_summary.setText(
            f"Angezeigt: {len(data)} Einträge | {len(groups)} Kategorie(n) | {total_all} Stück gesamt"
        )

    def _on_row_selected(self, current, _prev):
        if current is None:
            self._tbl_buchungen.setRowCount(0)
            self._lbl_detail_title.setText("Buchungshistorie  — Zeile auswählen")
            return
        # Andere Block-Tabellen deselektieren
        sender_tbl = self.sender()
        for tbl in self._block_tables:
            if tbl is not sender_tbl:
                tbl.blockSignals(True)
                tbl.clearSelection()
                tbl.setCurrentItem(None)
                tbl.blockSignals(False)
        item = current.data(Qt.ItemDataRole.UserRole)
        if not item:
            return
        art_id = item.get("art_id")
        groesse = item.get("groesse", "")
        art_name = item.get("art_name", "")
        self._lbl_detail_title.setText(f"Buchungshistorie: {art_name}  Gr. {groesse}")
        buchungen = self.db.get_buchungen_fuer_bestand(art_id, groesse)
        TYP_LABEL = {
            "eingang": "Eingang",
            "ausgabe": "Ausgabe",
            "rueckgabe": "Rückgabe",
            "entsorgung": "Entsorgung",
            "ausbuchen": "Ausbuchen",
            "korrektur": "Korrektur",
        }
        TYP_COLOR = {
            "eingang": QColor("#E8F5E9"),
            "ausgabe": QColor("#FFF3E0"),
            "rueckgabe": QColor("#E3F2FD"),
            "entsorgung": QColor("#FCE4EC"),
            "ausbuchen": QColor("#FCE4EC"),
            "korrektur": QColor("#F3E5F5"),
        }
        self._tbl_buchungen.setRowCount(len(buchungen))
        for r, b in enumerate(buchungen):
            typ = b.get("typ", "")
            cells = [
                format_datum(b.get("datum", "")),
                TYP_LABEL.get(typ, typ),
                str(b.get("menge", "")),
                b.get("mitarbeiter_name", "") or "",
                b.get("ausgegeben_von", "") or "",
                b.get("bemerkung", "") or "",
            ]
            color = TYP_COLOR.get(typ)
            for c, txt in enumerate(cells):
                cell = QTableWidgetItem(txt)
                cell.setFlags(cell.flags() & ~Qt.ItemFlag.ItemIsEditable)
                if color:
                    cell.setBackground(color)
                self._tbl_buchungen.setItem(r, c, cell)

    def _on_cell_double_clicked(self, row, _col):
        sender_tbl = self.sender()
        cell = sender_tbl.item(row, 0)
        if not cell:
            return
        item = cell.data(Qt.ItemDataRole.UserRole)
        if item:
            self._open_aktionen_dialog(item)

    def _open_aktionen_dialog(self, item: dict):
        dlg = QDialog(self)
        dlg.setWindowTitle(f"{item.get('art_name', '')}  Gr. {item.get('groesse', '')}")
        dlg.setMinimumWidth(320)
        dlg.setModal(True)
        layout = QVBoxLayout(dlg)
        layout.setSpacing(8)
        layout.setContentsMargins(20, 16, 20, 16)

        lbl = QLabel(f"<b>{item.get('art_name', '')}</b>  Gr. {item.get('groesse', '')}  –  Auf Lager: {item.get('menge', 0)}")
        layout.addWidget(lbl)

        result = [None]

        btn_eingang = QPushButton("Wareneingang buchen")
        btn_eingang.setObjectName("btn_secondary")
        def do_eingang():
            result[0] = "eingang"
            dlg.accept()
        btn_eingang.clicked.connect(do_eingang)
        layout.addWidget(btn_eingang)

        btn_edit = QPushButton("Bearbeiten  (Menge, Mindestbestand, Bemerkung)")
        btn_edit.setObjectName("btn_secondary")
        def do_edit():
            result[0] = "edit"
            dlg.accept()
        btn_edit.clicked.connect(do_edit)
        layout.addWidget(btn_edit)

        btn_del = QPushButton("Löschen")
        btn_del.setObjectName("btn_secondary")
        btn_del.setStyleSheet("color:#B20000;")
        def do_del():
            result[0] = "del"
            dlg.accept()
        btn_del.clicked.connect(do_del)
        layout.addWidget(btn_del)

        btn_cancel = QPushButton("Abbrechen")
        btn_cancel.clicked.connect(dlg.reject)
        layout.addWidget(btn_cancel)

        if dlg.exec() == QDialog.DialogCode.Accepted:
            if result[0] == "eingang":
                self._open_eingang_for(item)
            elif result[0] == "edit":
                self._open_edit(item)
            elif result[0] == "del":
                self._delete_item(item)

    def _open_add(self):
        art_id = self.cb_filter.currentData()
        self._open_add_for_art(art_id)

    def _open_add_for_art(self, art_id):
        dlg = BestandDialog(self.db, art_id=art_id, parent=self)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            d = dlg.get_data()
            self.db.upsert_bestand(
                d["art_id"], d["groesse"], d["menge"], d["min_menge"], d["bemerkung"]
            )
            self._load_data()

    def _open_edit(self, item: dict):
        dlg = BestandDialog(self.db, item=item, parent=self)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            d = dlg.get_data()
            self.db.upsert_bestand(
                item["art_id"], item["groesse"], d["menge"], d["min_menge"], d["bemerkung"]
            )
            self._load_data()

    def _open_eingang(self):
        art_id = self.cb_filter.currentData()
        self._open_eingang_for(None, art_id_preset=art_id)

    def _open_eingang_for(self, item, art_id_preset=None):
        art_id = item["art_id"] if item else art_id_preset
        groesse = str(item.get("groesse", "")) if item else ""
        dlg = EingangDialog(self.db, art_id=art_id, groesse=groesse, parent=self)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            errors = []
            ok_count = 0
            for d in dlg.get_data():
                ok, msg = self.db.eingang_kleidung(
                    d["art_id"], d["art_name"], d["groesse"],
                    d["menge"], d["datum"], d["bemerkung"]
                )
                if ok:
                    ok_count += 1
                else:
                    errors.append(msg)
            if ok_count:
                QMessageBox.information(self, "Eingang gebucht", f"{ok_count} Position(en) eingebucht.")
            if errors:
                QMessageBox.warning(self, "Fehler", "\n".join(errors))
            self._load_data()

    def _open_ausbuchen(self):
        dlg = MultiAusbuchenDialog(self.db, self)
        if dlg.exec() != QDialog.DialogCode.Accepted:
            return
        selected = dlg.get_selected()
        ok_count, errors = 0, []
        for entry in selected:
            item = entry["item"]
            ok, msg = self.db.ausbuchen_bestand(
                item["art_id"], item["art_name"], item["groesse"],
                entry["menge"], entry["datum"], entry["grund"], entry["bemerkung"]
            )
            if ok:
                ok_count += 1
            else:
                errors.append(msg)
        if ok_count:
            QMessageBox.information(self, "Ausgebucht", f"{ok_count} Position(en) ausgebucht.")
        if errors:
            QMessageBox.warning(self, "Fehler", "\n".join(errors))
        self._load_data()

    def _export_excel(self):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            QMessageBox.critical(self, "Fehler", "openpyxl nicht installiert. Bitte 'pip install openpyxl' ausführen.")
            return

        BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        export_dir = os.path.join(BASE, "Export")
        os.makedirs(export_dir, exist_ok=True)
        from datetime import date
        filename = os.path.join(export_dir, f"Bestand_{date.today().strftime('%Y%m%d')}.xlsx")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Bestand"

        header_fill = PatternFill("solid", fgColor="B20000")
        header_font = Font(bold=True, color="FFFFFF")
        headers = ["Kleidungsart", "Größe", "Auf Lager", "Mindestbestand", "Bemerkung"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(1, col, h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        warn_fill = PatternFill("solid", fgColor="FFF3CD")
        for r, item in enumerate(self._all_data, 2):
            menge = int(item.get("menge", 0))
            min_m = int(item.get("min_menge", 0))
            is_low = min_m > 0 and menge <= min_m
            row_data = [
                item.get("art_name", ""),
                str(item.get("groesse", "")),
                menge,
                min_m if min_m > 0 else "",
                item.get("bemerkung", ""),
            ]
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(r, col, val)
                if is_low:
                    cell.fill = warn_fill

        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        wb.save(filename)
        QMessageBox.information(self, "Excel-Export", f"Gespeichert:\n{filename}")

    def _delete_item(self, item: dict):
        reply = QMessageBox.question(
            self, "Eintrag löschen",
            f"Soll der Eintrag '{item['art_name']} Gr. {item['groesse']}' wirklich gelöscht werden?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
        )
        if reply == QMessageBox.StandardButton.Yes:
            self.db.delete_bestand_item(item["id"])
            self._load_data()

    def _open_neue_kategorie(self):
        dlg = QDialog(self)
        dlg.setWindowTitle("Neue Kleidungskategorie anlegen")
        dlg.setMinimumWidth(340)
        form = QFormLayout(dlg)
        form.setContentsMargins(20, 20, 20, 12)
        form.setSpacing(10)

        le_name = QLineEdit()
        le_name.setPlaceholderText("z.B. Weste, Mütze, Gürtel ...")
        le_beschr = QLineEdit()
        le_beschr.setPlaceholderText("Optional ...")

        form.addRow("Name *:", le_name)
        form.addRow("Beschreibung:", le_beschr)

        btns = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel
        )
        btns.button(QDialogButtonBox.StandardButton.Ok).setText("Anlegen")
        btns.button(QDialogButtonBox.StandardButton.Cancel).setText("Abbrechen")
        btns.accepted.connect(dlg.accept)
        btns.rejected.connect(dlg.reject)
        form.addRow(btns)

        if dlg.exec() != QDialog.DialogCode.Accepted:
            return

        name = le_name.text().strip()
        if not name:
            QMessageBox.warning(self, "Pflichtfeld", "Bitte einen Namen eingeben.")
            return

        try:
            self.db.add_kleidungsart(name, le_beschr.text().strip())
            QMessageBox.information(self, "Gespeichert", f'Kategorie „{name}“ wurde angelegt.')
            self._load_data()
        except Exception as e:
            QMessageBox.critical(self, "Fehler", str(e))

    def showEvent(self, event):
        super().showEvent(event)
        self._load_data()
