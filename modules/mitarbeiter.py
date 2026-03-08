"""
DRK Dienstkleidung - Mitarbeiterübersicht
Zeigt alle aktiven Mitarbeiter und ihre zugewiesene Kleidung.
"""

from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QSplitter, QLabel,
    QPushButton, QListWidget, QListWidgetItem, QTableWidget,
    QTableWidgetItem, QHeaderView, QAbstractItemView, QFrame,
    QLineEdit, QMessageBox, QDialog, QDialogButtonBox, QFormLayout,
)
from PySide6.QtCore import Qt, QSize
from PySide6.QtGui import QFont
import os

from utils import format_datum, export_table_to_csv
from modules.word_protokoll import create_bestand_protokoll, open_document


class MitarbeiterView(QWidget):
    """Mitarbeiter-Kleidungsübersicht."""

    PAGE_SIZE = 25

    def __init__(self, db, parent=None):
        super().__init__(parent)
        self.db = db
        self._current_ma_id: int | None = None
        self._current_ma_name: str = ""
        self._shown_count: int = self.PAGE_SIZE
        self._setup_ui()
        self._load_mitarbeiter()

    def _setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(24, 20, 24, 20)
        layout.setSpacing(10)

        # --- Titelzeile + Aktionen in einer Zeile ---
        top = QHBoxLayout()
        lbl_title = QLabel("Mitarbeiter-Kleidung")
        lbl_title.setObjectName("page_title")
        top.addWidget(lbl_title)
        top.addStretch()

        self.le_suche = QLineEdit()
        self.le_suche.setPlaceholderText("🔍  Name suchen ...")
        self.le_suche.setMaximumWidth(200)
        self.le_suche.textChanged.connect(self._filter_mitarbeiter)
        top.addWidget(self.le_suche)

        btn_reload_ma = QPushButton("↻")
        btn_reload_ma.setObjectName("btn_secondary")
        btn_reload_ma.setToolTip("Liste aktualisieren")
        btn_reload_ma.setFixedWidth(34)
        btn_reload_ma.clicked.connect(self._load_mitarbeiter)
        top.addWidget(btn_reload_ma)

        btn_new = QPushButton("➕  Neu anlegen")
        btn_new.setObjectName("btn_secondary")
        btn_new.clicked.connect(self._new_mitarbeiter)
        top.addWidget(btn_new)

        btn_del = QPushButton("🗑  Löschen")
        btn_del.setObjectName("btn_secondary")
        btn_del.setStyleSheet("color: #B20000;")
        btn_del.clicked.connect(self._delete_mitarbeiter)
        top.addWidget(btn_del)

        btn_excel_alle = QPushButton("📊  Excel-Export alle MA")
        btn_excel_alle.setObjectName("btn_secondary")
        btn_excel_alle.setToolTip("Kleidungsübersicht aller Mitarbeiter als Excel exportieren")
        btn_excel_alle.clicked.connect(self._export_alle_excel)
        top.addWidget(btn_excel_alle)

        layout.addLayout(top)

        # --- Splitter: links MA-Tabelle, rechts Kleidungsdetails ---
        splitter = QSplitter(Qt.Orientation.Horizontal)
        splitter.setChildrenCollapsible(False)

        # --- Linke Seite: MA-Tabelle + Mehr-Button + Zähler ---
        left = QWidget()
        left.setMinimumWidth(200)
        left.setMaximumWidth(320)
        left_layout = QVBoxLayout(left)
        left_layout.setContentsMargins(0, 0, 0, 0)
        left_layout.setSpacing(2)

        self._tbl_ma = QTableWidget(0, 2)
        self._tbl_ma.setHorizontalHeaderLabels(["Name", "Kleidung"])
        self._tbl_ma.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        self._tbl_ma.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        self._tbl_ma.verticalHeader().setVisible(False)
        self._tbl_ma.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self._tbl_ma.setAlternatingRowColors(True)
        self._tbl_ma.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self._tbl_ma.setSortingEnabled(True)
        self._tbl_ma.currentItemChanged.connect(self._on_ma_selected)
        left_layout.addWidget(self._tbl_ma)

        self._btn_mehr = QPushButton("▼  Weitere 25 laden")
        self._btn_mehr.setObjectName("btn_secondary")
        self._btn_mehr.clicked.connect(self._load_more_ma)
        self._btn_mehr.setVisible(False)
        left_layout.addWidget(self._btn_mehr)

        self._lbl_ma_count = QLabel("")
        self._lbl_ma_count.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self._lbl_ma_count.setStyleSheet("color: #888; font-size: 11px; padding: 2px;")
        left_layout.addWidget(self._lbl_ma_count)

        splitter.addWidget(left)

        # Rechte Seite: Detailansicht
        right = QWidget()
        right_layout = QVBoxLayout(right)
        right_layout.setContentsMargins(10, 0, 0, 0)
        right_layout.setSpacing(6)

        # Kompakter Header
        hdr_frame = QFrame()
        hdr_frame.setObjectName("stat_card")
        hdr_row = QHBoxLayout(hdr_frame)
        hdr_row.setContentsMargins(14, 8, 14, 8)
        hdr_row.setSpacing(10)

        self.lbl_ma_name = QLabel("Mitarbeiter auswählen")
        name_font = QFont()
        name_font.setBold(True)
        name_font.setPointSize(13)
        self.lbl_ma_name.setFont(name_font)
        self.lbl_ma_name.setStyleSheet("color: #B20000;")
        hdr_row.addWidget(self.lbl_ma_name)

        self.lbl_ma_info = QLabel("")
        self.lbl_ma_info.setObjectName("page_subtitle")
        hdr_row.addWidget(self.lbl_ma_info)
        hdr_row.addStretch()

        btn_reload = QPushButton("↻ Aktualisieren")
        btn_reload.setObjectName("btn_secondary")
        btn_reload.clicked.connect(self._reload_kleidung)
        hdr_row.addWidget(btn_reload)

        btn_protokoll = QPushButton("📄 Protokoll")
        btn_protokoll.setObjectName("btn_secondary")
        btn_protokoll.setToolTip("Kleidungsübersicht als Word-Protokoll erstellen")
        btn_protokoll.clicked.connect(self._protokoll_ma)
        hdr_row.addWidget(btn_protokoll)

        btn_export = QPushButton("📃 CSV-Export")
        btn_export.setObjectName("btn_secondary")
        btn_export.clicked.connect(self._export)
        hdr_row.addWidget(btn_export)

        right_layout.addWidget(hdr_frame)

        # Kleidungstabelle
        self._tbl = QTableWidget(0, 6)
        self._tbl.setHorizontalHeaderLabels(
            ["Kleidungsart", "Größe", "Anzahl", "Ausgabe-Datum", "Ausgeg. von", "Bemerkung"]
        )
        self._tbl.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        self._tbl.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        self._tbl.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        self._tbl.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)
        self._tbl.horizontalHeader().setSectionResizeMode(4, QHeaderView.ResizeMode.ResizeToContents)
        self._tbl.horizontalHeader().setSectionResizeMode(5, QHeaderView.ResizeMode.Stretch)
        self._tbl.verticalHeader().setVisible(False)
        self._tbl.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self._tbl.setAlternatingRowColors(True)
        self._tbl.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        right_layout.addWidget(self._tbl)

        self.lbl_gesamt = QLabel("")
        self.lbl_gesamt.setObjectName("page_subtitle")
        right_layout.addWidget(self.lbl_gesamt)

        splitter.addWidget(right)
        splitter.setSizes([260, 600])
        layout.addWidget(splitter)

    # ------------------------------------------------------------------

    def _load_more_ma(self):
        self._shown_count += self.PAGE_SIZE
        self._fill_ma_table(self._get_filtered_ma())

    def _get_filtered_ma(self) -> list[dict]:
        text = self.le_suche.text().lower()
        if text:
            return [e for e in self._all_ma_items if text in e["name"].lower()]
        return self._all_ma_items

    def _load_mitarbeiter(self):
        """Lade alle Mitarbeiter mit Kleidung in die linke Liste."""
        self._tbl_ma.setRowCount(0)
        self._shown_count = self.PAGE_SIZE
        self._all_ma_items: list[dict] = []

        # Mitarbeiter aus Kleidungsdatenbank (mit aktiver Ausgabe)
        ma_mit_kleidung = {
            mk["mitarbeiter_id"]: mk
            for mk in self.db.get_mitarbeiter_mit_kleidung()
        }

        # Aktive Mitarbeiter aus mitarbeiter.db
        alle_ma = self.db.get_alle_mitarbeiter()
        added_ids: set = set()

        for ma in alle_ma:
            mk_info = ma_mit_kleidung.get(ma["id"], {})
            entry = {
                "id": ma["id"],
                "name": f"{ma['nachname']}, {ma['vorname']}",
                "position": ma.get("position", ""),
                "abteilung": ma.get("abteilung", ""),
                "kleidung_anzahl": mk_info.get("anzahl_positionen", 0),
            }
            self._all_ma_items.append(entry)
            added_ids.add(ma["id"])

        # Historische Mitarbeiter (nur in kleidung.db vorhanden)
        for mk_id, mk_info in ma_mit_kleidung.items():
            if mk_id not in added_ids:
                entry = {
                    "id": mk_id,
                    "name": mk_info.get("mitarbeiter_name", f"ID {mk_id}"),
                    "position": "",
                    "abteilung": "",
                    "kleidung_anzahl": mk_info.get("anzahl_positionen", 0),
                }
                self._all_ma_items.append(entry)

        self._all_ma_items.sort(key=lambda x: x["name"])
        self._fill_ma_table(self._all_ma_items)

    def _fill_ma_table(self, items: list[dict]):
        shown = items[:self._shown_count]
        self._tbl_ma.setSortingEnabled(False)
        self._tbl_ma.setRowCount(len(shown))
        for r, entry in enumerate(shown):
            name_cell = QTableWidgetItem(entry["name"])
            name_cell.setData(Qt.ItemDataRole.UserRole, entry)
            if entry["kleidung_anzahl"] > 0:
                name_cell.setForeground(Qt.GlobalColor.darkGreen)
            self._tbl_ma.setItem(r, 0, name_cell)

            count_cell = QTableWidgetItem()
            count_cell.setData(Qt.ItemDataRole.DisplayRole, entry["kleidung_anzahl"])
            count_cell.setText(str(entry["kleidung_anzahl"]) if entry["kleidung_anzahl"] > 0 else "–")
            count_cell.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            if entry["kleidung_anzahl"] > 0:
                count_cell.setForeground(Qt.GlobalColor.darkGreen)
            self._tbl_ma.setItem(r, 1, count_cell)
        self._tbl_ma.setSortingEnabled(True)
        total = len(items)
        show_more = total > self._shown_count
        self._btn_mehr.setVisible(show_more)
        if show_more:
            remaining = total - self._shown_count
            self._btn_mehr.setText(f"▼  Weitere 25 laden ({remaining} verbleibend)")
        self._lbl_ma_count.setText(f"{min(self._shown_count, total)} von {total}")

    def _filter_mitarbeiter(self, text: str):
        self._shown_count = self.PAGE_SIZE
        self._fill_ma_table(self._get_filtered_ma())

    def _on_ma_selected(self, current, _prev):
        if current is None:
            return
        row = current.row()
        cell = self._tbl_ma.item(row, 0)
        if not cell:
            return
        entry = cell.data(Qt.ItemDataRole.UserRole)
        if not entry:
            return
        self._current_ma_id = entry["id"]
        self._current_ma_name = entry["name"]
        self._load_detail(entry)

    def _load_detail(self, entry: dict):
        self.lbl_ma_name.setText(entry["name"])
        info_parts = []
        if entry.get("position"):
            info_parts.append(entry["position"])
        if entry.get("abteilung"):
            info_parts.append(entry["abteilung"])
        self.lbl_ma_info.setText("  ·  ".join(info_parts) if info_parts else "")
        self._reload_kleidung()

    def _reload_kleidung(self):
        if self._current_ma_id is None:
            return

        items = self.db.get_mitarbeiter_kleidung(
            mitarbeiter_id=self._current_ma_id, status="ausgegeben"
        )
        if not items and self._current_ma_name:
            # Fallback: nach Name suchen (Einträge ohne ID-Zuordnung)
            # Anzeigename ist "Nachname, Vorname" → DB speichert "Vorname Nachname"
            display = self._current_ma_name.split("(")[0].strip()
            if ", " in display:
                nachname, vorname = display.split(", ", 1)
                name_vn = f"{vorname.strip()} {nachname.strip()}"   # "Sadi Tamer"
                name_nn = f"{nachname.strip()} {vorname.strip()}"   # "Tamer Sadi"
            else:
                name_vn = display
                name_nn = display
            for search_name in [name_vn, name_nn]:
                items = self.db.get_mitarbeiter_kleidung(
                    mitarbeiter_name=search_name, status="ausgegeben"
                )
                if items:
                    break

        self._tbl.setRowCount(len(items))
        gesamt = 0
        for r, it in enumerate(items):
            self._tbl.setItem(r, 0, QTableWidgetItem(it["art_name"]))
            self._tbl.setItem(r, 1, QTableWidgetItem(str(it["groesse"])))
            self._tbl.setItem(r, 2, QTableWidgetItem(str(it["menge"])))
            self._tbl.setItem(r, 3, QTableWidgetItem(format_datum(it["ausgabe_datum"])))
            self._tbl.setItem(r, 4, QTableWidgetItem(it.get("ausgegeben_von", "")))
            self._tbl.setItem(r, 5, QTableWidgetItem(it.get("bemerkung", "")))
            gesamt += int(it.get("menge", 0))
            for c in range(6):
                cell = self._tbl.item(r, c)
                if cell:
                    cell.setData(Qt.ItemDataRole.UserRole, it["id"])

        self.lbl_gesamt.setText(
            f"{len(items)} Position(en) · {gesamt} Stück gesamt" if items
            else "Keine aktive Kleidung zugeordnet"
        )

    def _new_mitarbeiter(self):
        dlg = QDialog(self)
        dlg.setWindowTitle("Neuen Mitarbeiter anlegen")
        dlg.setMinimumWidth(360)
        form = QFormLayout(dlg)
        form.setContentsMargins(20, 20, 20, 12)
        form.setSpacing(10)

        le_vn = QLineEdit()
        le_nn = QLineEdit()
        le_pnr = QLineEdit()
        le_pos = QLineEdit()
        le_abt = QLineEdit()

        form.addRow("Vorname *:", le_vn)
        form.addRow("Nachname *:", le_nn)
        form.addRow("Personalnummer:", le_pnr)
        form.addRow("Position:", le_pos)
        form.addRow("Abteilung:", le_abt)

        btns = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        btns.accepted.connect(dlg.accept)
        btns.rejected.connect(dlg.reject)
        form.addRow(btns)

        if dlg.exec() != QDialog.DialogCode.Accepted:
            return

        vn = le_vn.text().strip()
        nn = le_nn.text().strip()
        if not vn or not nn:
            QMessageBox.warning(self, "Pflichtfeld", "Vorname und Nachname sind Pflichtfelder.")
            return

        ok, msg = self.db.add_mitarbeiter(
            vn, nn, le_pnr.text().strip(),
            le_pos.text().strip(), le_abt.text().strip()
        )
        if ok:
            QMessageBox.information(self, "Gespeichert", msg)
            self._load_mitarbeiter()
        else:
            QMessageBox.critical(self, "Fehler", msg)

    def _delete_mitarbeiter(self):
        row = self._tbl_ma.currentRow()
        if row < 0:
            QMessageBox.information(self, "Hinweis", "Bitte zuerst einen Mitarbeiter in der Tabelle auswählen.")
            return
        cell = self._tbl_ma.item(row, 0)
        if not cell:
            return
        entry = cell.data(Qt.ItemDataRole.UserRole)
        if not entry or not entry.get("id"):
            QMessageBox.warning(self, "Nicht möglich", "Dieser Eintrag kann nicht gelöscht werden (kein Datenbankdatensatz).")
            return
        name = entry["name"]
        antwort = QMessageBox.question(
            self, "Mitarbeiter löschen",
            f"Mitarbeiter \u201e{name}\u201c wirklich dauerhaft aus der Datenbank löschen?\n\n"
            "Die zugeordnete Kleidungshistorie bleibt erhalten.",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.Cancel,
            QMessageBox.StandardButton.Cancel,
        )
        if antwort != QMessageBox.StandardButton.Yes:
            return

        ok, msg = self.db.delete_mitarbeiter(entry["id"])
        if ok:
            self._current_ma_id = None
            self._current_ma_name = ""
            self.lbl_ma_name.setText("Mitarbeiter auswählen")
            self.lbl_ma_info.setText("")
            self._tbl.setRowCount(0)
            self.lbl_gesamt.setText("")
            self._load_mitarbeiter()
        else:
            QMessageBox.critical(self, "Fehler", msg)

    def _export_alle_excel(self):
        """Exportiert Kleidungsübersicht aller Mitarbeiter als Excel."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            QMessageBox.critical(self, "Fehler", "openpyxl nicht installiert.\nBitte 'pip install openpyxl' ausführen.")
            return

        # Alle Mitarbeiter mit Kleidung laden
        ma_mit_kleidung = {
            mk["mitarbeiter_id"]: mk
            for mk in self.db.get_mitarbeiter_mit_kleidung()
        }
        alle_ma = self.db.get_alle_mitarbeiter()

        BASE = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
        export_dir = os.path.join(BASE, "Export")
        os.makedirs(export_dir, exist_ok=True)
        from datetime import date
        filename = os.path.join(export_dir, f"Mitarbeiter_Kleidung_{date.today().strftime('%Y%m%d')}.xlsx")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Mitarbeiter-Kleidung"

        header_fill = PatternFill("solid", fgColor="B20000")
        header_font = Font(bold=True, color="FFFFFF")
        headers = ["Name", "Position", "Abteilung", "Kleidungsart", "Größe", "Anzahl", "Ausgabe-Datum", "Ausgeg. von", "Bemerkung"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(1, col, h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        row_num = 2
        alt_fill = PatternFill("solid", fgColor="F5F5F5")

        for ma in sorted(alle_ma, key=lambda x: x.get("nachname", "")):
            ma_name = f"{ma['nachname']}, {ma['vorname']}"
            items = self.db.get_mitarbeiter_kleidung(mitarbeiter_id=ma["id"], status="ausgegeben")
            if items:
                first = True
                for it in items:
                    fill = alt_fill if (row_num % 2 == 0) else None
                    ws.cell(row_num, 1, ma_name if first else "").font = Font(bold=first)
                    if fill:
                        ws.cell(row_num, 1).fill = fill
                    ws.cell(row_num, 2, ma.get("position", "") if first else "")
                    ws.cell(row_num, 3, ma.get("abteilung", "") if first else "")
                    ws.cell(row_num, 4, it.get("art_name", ""))
                    ws.cell(row_num, 5, str(it.get("groesse", "")))
                    ws.cell(row_num, 6, int(it.get("menge", 0)))
                    ws.cell(row_num, 7, it.get("ausgabe_datum", ""))
                    ws.cell(row_num, 8, it.get("ausgegeben_von", "") or "")
                    ws.cell(row_num, 9, it.get("bemerkung", "") or "")
                    if fill:
                        for c in range(2, 10):
                            ws.cell(row_num, c).fill = fill
                    first = False
                    row_num += 1
            else:
                ws.cell(row_num, 1, ma_name)
                ws.cell(row_num, 1).font = Font(italic=True)
                ws.cell(row_num, 2, ma.get("position", ""))
                ws.cell(row_num, 3, ma.get("abteilung", ""))
                ws.cell(row_num, 4, "– keine Kleidung –")
                ws.cell(row_num, 4).font = Font(italic=True, color="888888")
                row_num += 1

        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 45)

        wb.save(filename)
        QMessageBox.information(self, "Excel-Export", f"Gespeichert:\n{filename}")

    def _protokoll_ma(self):
        """Erstellt ein Word-Protokoll der aktuellen Kleidung des Mitarbeiters."""
        if self._current_ma_id is None and not self._current_ma_name:
            QMessageBox.information(self, "Hinweis", "Bitte zuerst einen Mitarbeiter auswählen.")
            return
        if self._tbl.rowCount() == 0:
            QMessageBox.information(self, "Keine Kleidung", "Diesem Mitarbeiter ist keine Kleidung zugeordnet.")
            return

        from utils import today_iso
        artikel = []
        for r in range(self._tbl.rowCount()):
            mk_id = self._tbl.item(r, 0).data(Qt.ItemDataRole.UserRole) if self._tbl.item(r, 0) else None
            artikel.append({
                "art_name":      self._tbl.item(r, 0).text() if self._tbl.item(r, 0) else "",
                "groesse":       self._tbl.item(r, 1).text() if self._tbl.item(r, 1) else "",
                "menge":         self._tbl.item(r, 2).text() if self._tbl.item(r, 2) else "",
                "ausgabe_datum": self._tbl.item(r, 3).text() if self._tbl.item(r, 3) else "",
                "ausgegeben_von":self._tbl.item(r, 4).text() if self._tbl.item(r, 4) else "",
            })

        ok, path = create_bestand_protokoll(self._current_ma_name, today_iso(), artikel)
        if ok:
            open_document(path)
        else:
            QMessageBox.warning(self, "Protokoll-Fehler", path)

    def _export(self):
        headers = ["Mitarbeiter", "Kleidungsart", "Größe", "Anzahl", "Ausgabe-Datum", "Ausgegeben von", "Bemerkung"]
        rows = []
        name = self._current_ma_name or "Mitarbeiter"
        for r in range(self._tbl.rowCount()):
            row = [name]
            for c in range(6):
                it = self._tbl.item(r, c)
                row.append(it.text() if it else "")
            rows.append(row)

        ok, msg = export_table_to_csv(headers, rows, f"Kleidung_{name.replace(',', '').replace(' ', '_')}")
        if ok:
            QMessageBox.information(self, "Export erfolgreich", f"Datei gespeichert:\n{msg}")
        else:
            QMessageBox.warning(self, "Export fehlgeschlagen", msg)

    def showEvent(self, event):
        super().showEvent(event)
        self._load_mitarbeiter()
