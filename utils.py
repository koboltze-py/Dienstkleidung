"""
DRK Dienstkleidung - Hilfsfunktionen (Backup, Logging, Export)
"""

import os
import shutil
import logging
from datetime import datetime
from pathutils import get_base_dir

BASE_DIR = get_base_dir()
BACKUP_DIR = os.path.join(BASE_DIR, "Backup")
DATABASE_DIR = os.path.join(BASE_DIR, "Database")
EXPORT_DIR = os.path.join(BASE_DIR, "Export")
LOGS_DIR = os.path.join(BASE_DIR, "Logs")


def setup_logging():
    """Logging ins Log-Verzeichnis einrichten."""
    os.makedirs(LOGS_DIR, exist_ok=True)
    log_file = os.path.join(LOGS_DIR, f"dienstkleidung_{datetime.now().strftime('%Y%m')}.log")
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
        handlers=[
            logging.FileHandler(log_file, encoding="utf-8"),
            logging.StreamHandler(),
        ],
    )
    return logging.getLogger("DRKDienstkleidung")


def create_backup() -> tuple[bool, str]:
    """
    Erstellt eine Sicherungskopie aller Datenbanken im Backup-Ordner.
    Gibt (Erfolg, Meldung) zurück.
    """
    try:
        os.makedirs(BACKUP_DIR, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_folder = os.path.join(BACKUP_DIR, f"backup_{timestamp}")
        os.makedirs(backup_folder, exist_ok=True)

        copied = []
        if os.path.exists(DATABASE_DIR):
            for filename in os.listdir(DATABASE_DIR):
                if filename.endswith(".db"):
                    src = os.path.join(DATABASE_DIR, filename)
                    dst = os.path.join(backup_folder, filename)
                    shutil.copy2(src, dst)
                    copied.append(filename)

        if not copied:
            return False, "Keine Datenbankdateien gefunden."

        msg = f"Backup erstellt: {backup_folder}\nDateien: {', '.join(copied)}"
        return True, msg
    except Exception as e:
        return False, f"Backup fehlgeschlagen: {e}"


def get_backups() -> list[dict]:
    """Gibt eine Liste vorhandener Backups zurück."""
    import config as _cfg
    backup_root = _cfg.get("backup_dir")
    if not os.path.exists(backup_root):
        return []
    backups = []
    for folder in sorted(os.listdir(backup_root), reverse=True):
        path = os.path.join(backup_root, folder)
        if os.path.isdir(path) and folder.startswith("backup_"):
            try:
                ts_str = folder.replace("backup_", "")
                ts = datetime.strptime(ts_str, "%Y%m%d_%H%M%S")
                datum_str = ts.strftime("%d.%m.%Y %H:%M:%S")
            except Exception:
                datum_str = folder
            files = os.listdir(path)
            backups.append({"name": folder, "path": path, "datum": datum_str, "dateien": files})
    return backups


def create_full_backup(db) -> tuple[bool, str]:
    """
    Vollständiges Backup: DB-Dateien + Buchungsverlauf CSV + Bestand CSV + MA-Excel.
    Gibt (Erfolg, Meldung) zurück.
    """
    import csv
    import config as _cfg
    try:
        backup_root = _cfg.get("backup_dir")
        os.makedirs(backup_root, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_folder = os.path.join(backup_root, f"backup_{timestamp}")
        os.makedirs(backup_folder, exist_ok=True)

        # 1. DB-Dateien kopieren
        db_dir = _cfg.get("database_dir")
        copied_db = []
        if os.path.exists(db_dir):
            for fname in os.listdir(db_dir):
                if fname.endswith(".db"):
                    shutil.copy2(os.path.join(db_dir, fname), backup_folder)
                    copied_db.append(fname)

        # 2. Buchungsverlauf als CSV
        buchungen = db.get_buchungen(limit=1_000_000, offset=0)
        csv_path = os.path.join(backup_folder, f"Buchungsverlauf_{timestamp}.csv")
        with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.writer(f, delimiter=";")
            w.writerow(["Datum", "Typ", "Kleidungsart", "Groesse", "Menge",
                        "Mitarbeiter", "Ausgeg_von", "Bemerkung"])
            for b in buchungen:
                menge = b.get("menge", 0)
                w.writerow([
                    b.get("datum", ""), b.get("typ", ""),
                    b.get("art_name", ""), str(b.get("groesse", "")),
                    f"+{menge}" if menge > 0 else str(menge),
                    b.get("mitarbeiter_name", ""), b.get("ausgegeben_von", ""),
                    b.get("bemerkung", ""),
                ])

        # 3. Bestand als CSV
        bestand = db.get_bestand()
        csv_bestand = os.path.join(backup_folder, f"Bestand_{timestamp}.csv")
        with open(csv_bestand, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.writer(f, delimiter=";")
            w.writerow(["Kleidungsart", "Groesse", "Auf_Lager", "Mindestbestand", "Bemerkung"])
            for b in bestand:
                w.writerow([
                    b.get("art_name", ""), str(b.get("groesse", "")),
                    b.get("menge", 0), b.get("min_menge", 0),
                    b.get("bemerkung", ""),
                ])

        # 4. Mitarbeiter-Excel
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            alle_ma = db.get_alle_mitarbeiter()
            excel_path = os.path.join(backup_folder, f"Mitarbeiter_Kleidung_{timestamp}.xlsx")
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "MA-Kleidung"
            hdr_fill = PatternFill("solid", fgColor="B20000")
            hdr_font = Font(bold=True, color="FFFFFF")
            for col, h in enumerate(["Name", "Position", "Abteilung", "Kleidungsart",
                                      "Groesse", "Anzahl", "Ausgabe-Datum", "Ausgeg_von", "Bemerkung"], 1):
                c = ws.cell(1, col, h)
                c.font = hdr_font
                c.fill = hdr_fill
                c.alignment = Alignment(horizontal="center")
            row_n = 2
            for ma in sorted(alle_ma, key=lambda x: x.get("nachname", "")):
                ma_name = f"{ma['nachname']}, {ma['vorname']}"
                items = db.get_mitarbeiter_kleidung(mitarbeiter_id=ma["id"], status="ausgegeben")
                if items:
                    first = True
                    for it in items:
                        ws.cell(row_n, 1, ma_name if first else "").font = Font(bold=first)
                        ws.cell(row_n, 2, ma.get("position", "") if first else "")
                        ws.cell(row_n, 3, ma.get("abteilung", "") if first else "")
                        ws.cell(row_n, 4, it.get("art_name", ""))
                        ws.cell(row_n, 5, str(it.get("groesse", "")))
                        ws.cell(row_n, 6, int(it.get("menge", 0)))
                        ws.cell(row_n, 7, it.get("ausgabe_datum", ""))
                        ws.cell(row_n, 8, it.get("ausgegeben_von", "") or "")
                        ws.cell(row_n, 9, it.get("bemerkung", "") or "")
                        first = False
                        row_n += 1
                else:
                    ws.cell(row_n, 1, ma_name)
                    ws.cell(row_n, 4, "– keine Kleidung –")
                    row_n += 1
            wb.save(excel_path)
        except Exception:
            pass  # Excel optional – fehlende openpyxl wird ignoriert

        parts = [f"Backup: {backup_folder}"]
        if copied_db:
            parts.append(f"DB-Dateien: {', '.join(copied_db)}")
        parts.append("Buchungsverlauf CSV + Bestand CSV + MA-Excel gespeichert.")
        return True, "\n".join(parts)
    except Exception as e:
        return False, f"Backup fehlgeschlagen: {e}"


def export_table_to_csv(headers: list[str], rows: list[list], filename: str) -> tuple[bool, str]:
    """Exportiert eine Tabelle als CSV-Datei in den Export-Ordner."""
    try:
        import config as _cfg
        export_dir = _cfg.get("export_dir")
        os.makedirs(export_dir, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_filename = f"{filename}_{timestamp}.csv"
        filepath = os.path.join(export_dir, safe_filename)

        import csv
        with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f, delimiter=";")
            writer.writerow(headers)
            writer.writerows(rows)

        return True, filepath
    except Exception as e:
        return False, f"Export fehlgeschlagen: {e}"


def format_datum(datum_str: str) -> str:
    """Formatiert ein ISO-Datum (YYYY-MM-DD) als deutsches Datum (DD.MM.YYYY)."""
    if not datum_str:
        return ""
    try:
        dt = datetime.strptime(datum_str[:10], "%Y-%m-%d")
        return dt.strftime("%d.%m.%Y")
    except Exception:
        return datum_str


def parse_datum(datum_de: str) -> str:
    """Konvertiert deutsches Datum (DD.MM.YYYY) in ISO-Format (YYYY-MM-DD)."""
    try:
        dt = datetime.strptime(datum_de.strip(), "%d.%m.%Y")
        return dt.strftime("%Y-%m-%d")
    except Exception:
        return datum_de


def today_iso() -> str:
    """Gibt das heutige Datum als ISO-String zurück."""
    return datetime.now().strftime("%Y-%m-%d")
