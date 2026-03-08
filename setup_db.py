"""
DRK Dienstkleidung - setup_db.py
Einmalige Initialisierung: Liest Excel-Datei und erstellt kleidung.db
Ausführen mit: python setup_db.py
"""

import sqlite3
import os
import re
import sys
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("Fehler: openpyxl ist nicht installiert. Bitte 'pip install openpyxl' ausführen.")
    sys.exit(1)

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
KLEIDUNG_DB     = os.path.join(BASE_DIR, "Database", "kleidung.db")
MITARBEITER_DB  = os.path.join(BASE_DIR, "Database", "mitarbeiter.db")
EXCEL_PATH = os.path.join(BASE_DIR, "Data", "Bestand Excel", "Kleiderkammer aktuell 01.12.25.xlsx")

# Kleidungsarten und ihre Spaltenpositionen in der Bestandstabelle (0-basiert)
BESTAND_KATEGORIEN = [
    {"name": "Diensthosen", "g_col": 1,  "m_col": 2},   # B, C
    {"name": "Schuhe",      "g_col": 5,  "m_col": 6},   # F, G
    {"name": "Pullover",    "g_col": 9,  "m_col": 10},  # J, K
    {"name": "T-Shirts",    "g_col": 13, "m_col": 14},  # N, O
    {"name": "Jacken",      "g_col": 17, "m_col": 18},  # R, S
]

# Spalten in der Ausgabeliste (0-basiert, ab Zeile 29 im Excel)
AUSGABE_KATEGORIEN = [
    {"name": "Schuhe",      "g_col": 5,  "m_col": 6},   # F, G
    {"name": "Pullover",    "g_col": 9,  "m_col": 10},  # J, K
    {"name": "T-Shirts",    "g_col": 13, "m_col": 14},  # N, O
    {"name": "Jacken",      "g_col": 17, "m_col": 18},  # R, S
    {"name": "Diensthosen", "g_col": 20, "m_col": 21},  # U, V
]

BESTAND_START = 3    # 1-basierte Excel-Zeilennummer, bis Zeile 26
BESTAND_ENDE  = 26
AUSGABE_KOPF  = 29   # Zeile mit Spaltenkopf "Datum, Name, ..."
AUSGABE_START = 30   # Erste Datenzelle

CREATE_SQL = """
PRAGMA foreign_keys = ON;

CREATE TABLE IF NOT EXISTS kleidungsarten (
    id           INTEGER PRIMARY KEY AUTOINCREMENT,
    name         TEXT    NOT NULL UNIQUE,
    beschreibung TEXT    DEFAULT '',
    aktiv        INTEGER DEFAULT 1,
    erstellt_am  TEXT    DEFAULT (datetime('now','localtime'))
);

CREATE TABLE IF NOT EXISTS kleidungsbestand (
    id           INTEGER PRIMARY KEY AUTOINCREMENT,
    art_id       INTEGER NOT NULL REFERENCES kleidungsarten(id) ON DELETE CASCADE,
    groesse      TEXT    NOT NULL,
    menge        INTEGER NOT NULL DEFAULT 0,
    min_menge    INTEGER DEFAULT 0,
    bemerkung    TEXT    DEFAULT '',
    erstellt_am  TEXT    DEFAULT (datetime('now','localtime')),
    geaendert_am TEXT    DEFAULT (datetime('now','localtime')),
    UNIQUE(art_id, groesse)
);

CREATE TABLE IF NOT EXISTS mitarbeiter_kleidung (
    id               INTEGER PRIMARY KEY AUTOINCREMENT,
    mitarbeiter_id   INTEGER,
    mitarbeiter_name TEXT    NOT NULL,
    art_id           INTEGER NOT NULL REFERENCES kleidungsarten(id),
    art_name         TEXT    NOT NULL,
    groesse          TEXT    NOT NULL,
    menge            INTEGER NOT NULL DEFAULT 1,
    ausgabe_datum    TEXT    NOT NULL,
    rueckgabe_datum  TEXT,
    status           TEXT    DEFAULT 'ausgegeben',
    ausgegeben_von   TEXT    DEFAULT '',
    bemerkung        TEXT    DEFAULT '',
    erstellt_am      TEXT    DEFAULT (datetime('now','localtime'))
);

CREATE TABLE IF NOT EXISTS buchungen (
    id               INTEGER PRIMARY KEY AUTOINCREMENT,
    art_id           INTEGER,
    art_name         TEXT    NOT NULL,
    groesse          TEXT    NOT NULL,
    menge            INTEGER NOT NULL,
    typ              TEXT    NOT NULL,
    mitarbeiter_id   INTEGER,
    mitarbeiter_name TEXT    DEFAULT '',
    datum            TEXT    NOT NULL,
    ausgegeben_von   TEXT    DEFAULT '',
    bemerkung        TEXT    DEFAULT '',
    erstellt_am      TEXT    DEFAULT (datetime('now','localtime'))
);

CREATE INDEX IF NOT EXISTS idx_bestand_art     ON kleidungsbestand(art_id);
CREATE INDEX IF NOT EXISTS idx_buchungen_datum ON buchungen(datum);
CREATE INDEX IF NOT EXISTS idx_buchungen_ma    ON buchungen(mitarbeiter_id);
CREATE INDEX IF NOT EXISTS idx_mk_mitarbeiter  ON mitarbeiter_kleidung(mitarbeiter_id);
CREATE INDEX IF NOT EXISTS idx_mk_name         ON mitarbeiter_kleidung(mitarbeiter_name);
CREATE INDEX IF NOT EXISTS idx_mk_status       ON mitarbeiter_kleidung(status);
"""


def _parse_menge(wert) -> int:
    """Extrahiert eine integer-Menge aus einem Zellwert."""
    if wert is None:
        return 0
    if isinstance(wert, (int, float)):
        try:
            return max(0, int(wert))
        except Exception:
            return 0
    match = re.search(r"\d+", str(wert))
    return int(match.group()) if match else 0


def _clean_str(wert) -> str:
    """Gibt einen sauberen String zurück."""
    if wert is None:
        return ""
    s = str(wert).strip()
    return "" if s.lower() == "none" else s


def setup():
    # Sicherheitsabfrage
    if os.path.exists(KLEIDUNG_DB):
        antwort = input(
            f"\nDatenbank '{KLEIDUNG_DB}' existiert bereits.\n"
            "Überschreiben und neu erstellen? (j/n): "
        ).strip().lower()
        if antwort != "j":
            print("Abgebrochen.")
            return
        os.remove(KLEIDUNG_DB)
        print("Alte Datenbank gelöscht.")

    if not os.path.exists(EXCEL_PATH):
        print(f"FEHLER: Excel-Datei nicht gefunden:\n{EXCEL_PATH}")
        sys.exit(1)

    # Datenbank erstellen
    print("\n[1/4] Erstelle Datenbanktabellen...")
    conn = sqlite3.connect(KLEIDUNG_DB)
    conn.executescript(CREATE_SQL)
    conn.commit()

    # Kleidungsarten einfügen
    print("[2/4] Füge Kleidungsarten ein...")
    art_ids: dict[str, int] = {}
    for kat in BESTAND_KATEGORIEN:
        cur = conn.execute(
            "INSERT OR IGNORE INTO kleidungsarten (name) VALUES (?)", (kat["name"],)
        )
        if cur.lastrowid:
            art_ids[kat["name"]] = cur.lastrowid
        else:
            art_ids[kat["name"]] = conn.execute(
                "SELECT id FROM kleidungsarten WHERE name=?", (kat["name"],)
            ).fetchone()[0]
    conn.commit()
    print(f"   {len(art_ids)} Kleidungsarten angelegt: {', '.join(art_ids)}")

    # Mitarbeiter-IDs aus mitarbeiter.db laden (für korrekte Zuordnung)
    ma_id_lookup: dict[str, int] = {}  # key: name in Kleinbuchstaben → id
    if os.path.exists(MITARBEITER_DB):
        ma_conn = sqlite3.connect(MITARBEITER_DB)
        ma_conn.row_factory = sqlite3.Row
        for row in ma_conn.execute("SELECT id, vorname, nachname FROM mitarbeiter"):
            vn = (row["vorname"] or "").strip()
            nn = (row["nachname"] or "").strip()
            if vn and nn:
                # Beide Reihenfolgen speichern ("Vorname Nachname" + "Nachname Vorname")
                ma_id_lookup[f"{vn} {nn}".lower()] = row["id"]
                ma_id_lookup[f"{nn} {vn}".lower()] = row["id"]
            elif vn:
                ma_id_lookup[vn.lower()] = row["id"]
            elif nn:
                ma_id_lookup[nn.lower()] = row["id"]
        ma_conn.close()
        print(f"   {len(ma_id_lookup) // 2} Mitarbeiter für ID-Zuordnung geladen.")
    else:
        print("   WARNUNG: mitarbeiter.db nicht gefunden – mitarbeiter_id bleibt NULL.")

    # Excel-Datei lesen
    print(f"[3/4] Lese Excel-Datei:\n   {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb["Tabelle1"]
    alle_zeilen = list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True))

    # --- Bestandsdaten importieren ---
    print("     Importiere Bestandsdaten (Zeilen %d-%d)..." % (BESTAND_START, BESTAND_ENDE))
    bestand_count = 0
    for zeil_idx in range(BESTAND_START - 1, BESTAND_ENDE):
        zeile = alle_zeilen[zeil_idx]
        for kat in BESTAND_KATEGORIEN:
            groesse_raw = zeile[kat["g_col"]] if kat["g_col"] < len(zeile) else None
            menge_raw   = zeile[kat["m_col"]] if kat["m_col"] < len(zeile) else None

            groesse = _clean_str(groesse_raw)
            if not groesse:
                continue

            menge = _parse_menge(menge_raw)
            art_id = art_ids[kat["name"]]

            conn.execute(
                """INSERT INTO kleidungsbestand (art_id, groesse, menge)
                   VALUES (?, ?, ?)
                   ON CONFLICT(art_id, groesse) DO UPDATE SET
                     menge = menge + excluded.menge""",
                (art_id, groesse, menge),
            )
            bestand_count += 1

    conn.commit()
    print(f"     {bestand_count} Bestandseinträge importiert.")

    # --- Ausgabehistorie importieren ---
    print("     Importiere Ausgabehistorie (ab Zeile %d)..." % AUSGABE_START)
    ausgabe_count = 0
    for zeil_idx in range(AUSGABE_START - 1, len(alle_zeilen)):
        zeile = alle_zeilen[zeil_idx]

        datum_raw  = zeile[0] if len(zeile) > 0 else None
        nachname   = zeile[1] if len(zeile) > 1 else None
        vorname    = zeile[2] if len(zeile) > 2 else None
        ausg_von   = zeile[23] if len(zeile) > 23 else None
        notiz      = zeile[24] if len(zeile) > 24 else None

        if datum_raw is None and nachname is None:
            continue

        # Datum auflösen
        if isinstance(datum_raw, datetime):
            datum = datum_raw.strftime("%Y-%m-%d")
        else:
            d = _clean_str(datum_raw)
            # Versuche deutsches Format
            datum = "2025-12-01"
            for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d.%m.%y"):
                try:
                    datum = datetime.strptime(d, fmt).strftime("%Y-%m-%d")
                    break
                except Exception:
                    pass

        # Mitarbeitername
        nachname_s = _clean_str(nachname)
        if not nachname_s:
            continue
        vorname_s = _clean_str(vorname)
        ma_name = f"{vorname_s} {nachname_s}".strip() if vorname_s else nachname_s

        ausg_von_s = _clean_str(ausg_von)
        notiz_s    = _clean_str(notiz)

        # Mitarbeiter-ID nachschlagen (versuche beide Namensreihenfolgen)
        ma_id = ma_id_lookup.get(ma_name.lower())

        for kat in AUSGABE_KATEGORIEN:
            groesse_raw = zeile[kat["g_col"]] if kat["g_col"] < len(zeile) else None
            menge_raw   = zeile[kat["m_col"]] if kat["m_col"] < len(zeile) else None

            groesse = _clean_str(groesse_raw)
            if not groesse:
                continue

            menge = max(1, _parse_menge(menge_raw)) if menge_raw is not None else 1
            art_id = art_ids.get(kat["name"])
            if art_id is None:
                continue

            conn.execute(
                """INSERT INTO buchungen
                   (art_id, art_name, groesse, menge, typ, mitarbeiter_id,
                    mitarbeiter_name, datum, ausgegeben_von, bemerkung)
                   VALUES (?, ?, ?, ?, 'ausgabe', ?, ?, ?, ?, ?)""",
                (art_id, kat["name"], groesse, -menge, ma_id,
                 ma_name, datum, ausg_von_s, notiz_s),
            )
            conn.execute(
                """INSERT INTO mitarbeiter_kleidung
                   (mitarbeiter_id, mitarbeiter_name, art_id, art_name, groesse, menge,
                    ausgabe_datum, ausgegeben_von, bemerkung, status)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 'ausgegeben')""",
                (ma_id, ma_name, art_id, kat["name"], groesse, menge,
                 datum, ausg_von_s, notiz_s),
            )
            ausgabe_count += 1

    conn.commit()
    conn.close()
    print(f"     {ausgabe_count} Ausgabeeinträge aus Historieblatt importiert.")

    print("\n[4/4] Fertig!")
    print(f"   Datenbank: {KLEIDUNG_DB}")
    print(f"   Kleidungsarten:     {len(art_ids)}")
    print(f"   Bestandseinträge:   {bestand_count}")
    print(f"   Ausgabeeinträge:    {ausgabe_count}")
    print("\nDie App kann jetzt mit 'python main.py' gestartet werden.")


if __name__ == "__main__":
    setup()
